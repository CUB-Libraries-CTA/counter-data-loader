import openpyxl
import collections
import os, csv
from datetime import datetime

class TitleMasterReport:
    """
    Represents a Title Master Report.

    This class is able to load any of the standard views defined for the
    Title Master Report. Currently, only J1/J3 reports are loaded for
    journal titles and the B1/B3 reports for books. The type of standard view
    is specified in the header (Report_ID).
    """

    PERIODS = ['', 'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug',
        'sep', 'oct', 'nov', 'dec']
    ACCESS_TYPE = {'Controlled': 1, 'OA_Gold': 2, 'Other_Free_To_Read': 3}
    METRIC_TYPE = {'Total_Item_Investigations': 1, 'Total_Item_Requests': 2, 'Unique_Item_Investigations': 3,
        'Unique_Item_Requests': 4, 'Unique_Title_Investigations': 5, 'Unique_Title_Requests': 6,
        'Limit_Exceeded': 7, 'No_License': 8}
    MAX_COLS = 16384
    MAX_ROWS = 1048576
    HEADER_ROW = 14
    DATA_ROW_START = 15
    DATA_COL_START = 1

    def __init__(self, workbook):
        self._workbook = openpyxl.load_workbook(filename=workbook, data_only=True, read_only=True)
        self._worksheet = self._workbook.active
        self._report_id = self._worksheet.cell(row=2, column=2).value
        self._reporting_period = self._worksheet.cell(row=10, column=2).value
        self._run_date = self._worksheet.cell(row=11, column=2).value
        self._platform = self._worksheet.cell(row=15, column=4).value
        self._filename = os.path.basename(workbook)

    @property
    def filename(self):
        return self._filename

    @property
    def report_id(self):
        return self._report_id

    @property
    def begin_date(self):
        kv_pair = self._reporting_period.split(';')[0]
        return kv_pair.split('=')[1]

    @property
    def end_date(self):
        kv_pair = self._reporting_period.split(';')[1]
        return kv_pair.split('=')[1]
    
    @property
    def run_date(self):
        if self._run_date is None:
            return '0000-00-00'
        else:
            return str(self._run_date)[0:10]

    @property
    def title_type(self):
        return self.report_id[3:4]
    
    @property
    def platform(self):
        return self._platform

    @property
    def row_count(self):
        return len(self.data_rows())
    
    def close(self):
        self._workbook.close()
    
    def _header_row(self):
        """
        Returns the report header row (column names).

        The header row is constructed from two pieces:

            - all the title identifying information (incl isbn and yop)
            - the usage period columns

        Access_Type may be included in either report type, e.g., in TR_B3 and
        TR_J3.

        For the date columns, the exact number (and range) depends on the
        reporting period of the report. In most cases, the reporting period
        will be January to December; however, the report could have been
        run for any period, e.g., April to July. Therefore, the begin and
        end dates for the reporting period will determine which month columns
        to include in the header.
        """

        # Build the title/book info columns by iterating through the spreadsheet
        # columns beginning with A (Title)
        header = ['report_id', 'title_type', 'title', 'publisher', 'publisher_id',
            'platform', 'doi', 'proprietary_id', 'isbn', 'print_issn', 'online_issn',
            'uri', 'yop', 'access_type', 'metric_type']
        start_month = int(self.begin_date[5:7])
        end_month = int(self.end_date[5:7])
        period_cols = self.PERIODS[start_month:end_month+1]
        for month in period_cols:
            header.append(month)

        return header

    def data_rows(self):
        """
        Returns the range of data rows. For all report types, this is rows
        15 and onwards. The actual number of rows in a given report is
        variable and depends on the publisher.
        """

        n = 0
        for row in self._worksheet.iter_rows(min_row=self.DATA_ROW_START, min_col=1,
            max_row=self.MAX_ROWS, max_col=1, values_only=True):
            if row[0] is None: # When no more data, the first cell in the row will be blank
                break
            n += 1

        return range(self.DATA_ROW_START, self.DATA_ROW_START + n)

    def _data_cols(self):
        """
        Returns the range of data columns. The actual number of columns
        depends on the report type.
        """

        n = 0
        header_row = self._worksheet[self.HEADER_ROW]
        for cell in header_row:
            if cell.value is None:
                break
            n += 1

        return range(self.DATA_COL_START, self.DATA_COL_START + n)

    def get_row(self, n):
        """
        Gets a data row from the spreadsheet for a given row number.

        The return value is a named tuple, which provides for more
        intuitive referencing of columns during inserts and updates.
        """

        row = self._worksheet[n]
        row_spec = collections.namedtuple('ReportRow', self._header_row())

        # Initialize the data row
        datarow = list()
        datarow.append(self.report_id)
        datarow.append(self.title_type)

        # Append the remaining column data
        i = 0
        while i < len(row):
            if i == 6 and self.title_type == 'J':
                datarow.append('') # isbn
            if i == 9 and self.title_type == 'J':
                datarow.append('') # yop
            if i == 9 and self.report_id == 'TR_J1':
                datarow.append('Controlled') # access_type
            if row[i].value is None:
                datarow.append('')
            else:
                datarow.append(str(row[i].value).strip())
            i += 1

        # Remove the total columns that are not used
        datarow = datarow[0:15] + datarow[16:]

        return row_spec._make(datarow)

    def export(self, dir):
        """
        Makes text files of the raw spreadsheet data for bulk import into the DB.
        """
        # Start with titles
        with open('{0}/title_report_temp'.format(dir), 'w', newline='', encoding='utf-8') as csvfile:
            csvwriter = csv.writer(csvfile, dialect='excel-tab', lineterminator='\n')
        
            # Iterate through the spreadsheet rows starting at the first data row (row 10).
            datarows = self.data_rows()
            row_num = min(datarows)
            for row in self._worksheet.iter_rows(min_row=min(datarows), min_col=1, max_row=max(datarows), max_col=11):
                # Start building a list of field values. The actual fields and their sequence
                # must correspond to the title_report_temp table. See schema for details.
                datarow = list()
                datarow.append('null') # id
                i = 0
                while i < len(row):
                    if i == 1:
                        datarow.append(self.title_type) # title_type
                    if i == 6 and self.title_type == 'J':
                        datarow.append('') # isbn
                    if i == 9 and self.title_type == 'J':
                        datarow.append('') # yop
                        break
                    if row[i].value is None:
                        datarow.append('')
                    else:
                        datarow.append(str(row[i].value).strip())
                    i += 1
                datarow.append(self.filename) # excel_name
                datarow.append(row_num) # row_num
                datarow.append(0) # title_report_id
                csvwriter.writerow(datarow)
                row_num += 1


        # Export metrics
        with open('{0}/metric_temp'.format(dir), 'w', newline='', encoding='utf-8') as csvfile:
            csvwriter = csv.writer(csvfile, dialect='excel-tab', lineterminator='\n')

            report_begin = datetime.fromisoformat(self.begin_date)
            report_end = datetime.fromisoformat(self.end_date)

            # Iterate through the spreadsheet rows starting at the first data row (row 15).
            # Period columns either start at J(10) for journals or L(12) for books.
            min_col = {'J': 10, 'B': 12}
            datarows = self.data_rows()
            row_num = min(datarows)
            for row in self._worksheet.iter_rows(min_row=min(datarows), min_col=min_col[self.title_type],
                max_row=max(datarows), max_col=len(self._data_cols())):

                # A row will be inserted for each month column in the source
                # spreadsheet. The actual number of months is determined from
                # the start and end dates contained in the report header.
                n = 3
                for i in range(report_begin.month, report_end.month + 1):
                    period = datetime(report_begin.year, i, 1).strftime('%Y-%m-%d')
                    period_total = int(float(row[n].value)) # float conversion deals with cases of '0.0'
                    access_type = self.ACCESS_TYPE[str(row[0].value.strip())]
                    metric_type = self.METRIC_TYPE[str(row[1].value.strip())]

                    # Start building a list of field values. The actual fields and their sequence
                    # must correspond to the title_report_temp table. See schema for details.
                    datarow = list()
                    datarow.append('null') # id
                    datarow.append(0) # title_report_id
                    datarow.append(access_type) # access_type
                    datarow.append(metric_type) # metric_type
                    datarow.append(period)
                    datarow.append(period_total)
                    datarow.append(self.filename)
                    datarow.append(row_num)
                    csvwriter.writerow(datarow)
                    n += 1
                row_num += 1
