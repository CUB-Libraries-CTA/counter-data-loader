import openpyxl
import collections
import os, csv
from datetime import datetime

class JR1Report:
    """
    Represents the JR1 report spreadsheet.

    The JR1 report is defined by the COUNTER Code of Practice R4 specification.
    The layout and data requirements are very specific. The properties and methods
    defined in this class rely on adherence to this specification otherwise data
    anomalies may arise and/or code using this class may raise an exception.

    NOTE: JR1 reports are no longer used with the COUNTER R5 specification, which the CU
    Libraries started using exclusively in 2021.   So JR1 reports no longer need to be
    supported.
    """

    PERIODS = ['', 'jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug',
        'sep', 'oct', 'nov', 'dec']
    MAX_ROWS = 1048576
    DATA_ROW_START = 10

    def __init__(self, workbook):
        self._workbook = openpyxl.load_workbook(filename=workbook, data_only=True, read_only=True)
        self._worksheet = self._workbook.active
        self._report_id = self._worksheet.cell(row=1, column=1).value
        self._reporting_period = self._worksheet.cell(row=5, column=1).value
        self._run_date = self._worksheet.cell(row=7, column=1).value
        self._platform = self._worksheet.cell(row=10, column=3).value
        self._filename = os.path.basename(workbook)
        self._dirname = os.path.dirname(workbook)
    
    @property
    def filename(self):
        return self._filename
    
    @property
    def report_id(self):
        return self._report_id
    
    @property
    def begin_date(self):
        return self._reporting_period.split('to')[0].strip()

    @property
    def end_date(self):
        return self._reporting_period.split('to')[1].strip()
    
    @property
    def run_date(self):
        if self._run_date is None:
            return '0000-00-00'
        else:
            return str(self._run_date)[0:10]
    
    @property
    def title_type(self):
        return 'J'
    
    @property
    def platform(self):
        return self._platform
    
    @property
    def row_count(self):
        return len(self.data_rows())
    
    def close(self):
        self._workbook.close()

    def _header_row(self):
        """
        Returns the header row

        The header row for J1 (R4) report consists of 12 static
        columns and a variable number of date columns, depending
        on the reporting period. For a 12 month report, there will
        be 12 date columns (Jan - Dec).

        To force compatibility with the R5 specification, additional
        columns are introduced and are populated with empty strings
        or predetermined values that are consistent for all JR1 reports.

        Note that the first two columns are not part of the report
        specification. They are included for later referencing
        when data is loaded into the database.

        DEPRECATED when using bulk import method.
        """

        # Start with non-date columns, which are static, and then
        # append the date column headings
        header = ['report_id', 'title_type', 'title', 'publisher', 'publisher_id',
            'platform', 'doi', 'proprietary_id', 'isbn', 'print_issn', 'online_issn',
            'uri', 'yop', 'access_type', 'metric_type']
        start_month = int(self.begin_date.split('-')[1])
        end_month = int(self.end_date.split('-')[1])
        period_cols = self.PERIODS[start_month:end_month+1]
        for month in period_cols:
            header.append(month)

        return header

    def data_rows(self):
        """
        Returns the number of rows containing publication data.

        According to the COUNTER Code of Practice R4, publication data always
        starts at row 10 in the spreadsheet. Assuming a continuous list of
        data rows, the last row will be followed by a blank row.
        """

        n = 0
        for row in self._worksheet.iter_rows(min_row=self.DATA_ROW_START, min_col=1,
            max_row=self.MAX_ROWS, max_col=1, values_only=True):
            if row[0] is None: # Done when the first cell in the row is blank
                break
            n += 1
            
        return range(self.DATA_ROW_START, self.DATA_ROW_START + n)

    def get_row(self, n):
        """
        Returns a complete row of publication data for the given row number.

        For a JR1 report covering twelve months of usage data (typically Jan-Dec),
        there will be 22 columns of data in the row. Reports containing less (or more)
        months of usage data will have less or more list items, respectively.

        DEPRECATED when using bulk import method.
        """

        row = self._worksheet[n]
        row_spec = collections.namedtuple('ReportRow', self._header_row())

        # Initialize the data row
        datarow = list()
        datarow.append(self.report_id)
        datarow.append(self.title_type)

        # Append the remaining column data
        i = 0
        while i < len(row):
            if i in (2, 5):
                datarow.append('') # publisher_id and isbn
            if i == 7:
                datarow.append('') # uri
                datarow.append('') # yop
                datarow.append('Controlled') # access_type 
                datarow.append('Total_Item_Requests') # metric_type
            if row[i].value is None:
                datarow.append('')
            else:
                datarow.append(str(row[i].value).strip())
            i += 1
        
        # Remove the total columns that are not used
        datarow = datarow[0:15] + datarow[18:]

        return row_spec._make(datarow)
    
    def export(self):
        """
        Prepares text files of the raw spreadsheet data for bulk import into the DB. Two separate
        files are generated (titles and metrics):

        - title_report_temp
        - metric_temp
        
        Both files are Excel tab delimited format, which are subsequently loaded with mysqlimport.
        """
        
        # Start with titles data.
        title_report_temp = '{0}/title_report_temp'.format(self._dirname)
        with open(title_report_temp, 'w', newline='', encoding='utf-8') as csvfile:
            csvwriter = csv.writer(csvfile, dialect='excel-tab', lineterminator='\n')
        
            # Iterate through the spreadsheet rows starting at the first data row (row 10). Only the
            # first 7 columns are included in the export for titles data.
            datarows = self.data_rows() # Range of rows in the spreadsheet.
            row_num = min(datarows) # Spreadsheet row number.
            for row in self._worksheet.iter_rows(min_row=min(datarows), min_col=1, max_row=max(datarows), max_col=7):
                # Start building a list of field values. The actual fields and their sequence
                # must correspond to the title_report_temp table. See schema for details.
                datarow = list()
                datarow.append('null') # id
                i = 0
                while i <= len(row):
                    if i == 1:
                        datarow.append(self.title_type) # title_type
                    if i == 2:
                        datarow.append('') # publisher_id
                    if i == 5:
                        datarow.append('') # isbn
                    if i == 7:
                        datarow.append('') # uri
                        datarow.append('') # yop
                        break
                    if row[i].value is None:
                        datarow.append('') # cell is blank
                    else:
                        datarow.append(str(row[i].value).strip())
                    i += 1
                datarow.append(self._filename) # excel_name
                datarow.append(row_num) # row_num
                datarow.append(0) # title_report_id
                csvwriter.writerow(datarow)
                row_num += 1

        # Export metrics.
        metric_temp = '{0}/metric_temp'.format(self._dirname)
        with open(metric_temp, 'w', newline='', encoding='utf-8') as csvfile:
            csvwriter = csv.writer(csvfile, dialect='excel-tab', lineterminator='\n')

            # Grab the report begin and end dates, which are needed to generate the report periods.
            report_begin = datetime.fromisoformat(self.begin_date)
            report_end = datetime.fromisoformat(self.end_date)

            # Iterate through the spreadsheet rows starting at the first data row (row 10). For metrics,
            # the first applicable report column is 11 and extends to the number of months in the report.
            datarows = self.data_rows()
            row_num = min(datarows)
            for row in self._worksheet.iter_rows(min_row=min(datarows), min_col=11, max_row=max(datarows), max_col=report_end.month+10):
                # A row will be inserted for each month column in the source
                # spreadsheet. The actual number of months is determined from
                # the start and end dates contained in the report header.
                n = 0
                for i in range(report_begin.month, report_end.month + 1):
                    period = datetime(report_begin.year, i, 1).strftime('%Y-%m-%d')
                    period_total = int(float(row[n].value)) # float conversion deals with cases of '0.0'

                    # Start building a list of field values. The actual fields and their sequence
                    # must correspond to the title_report_temp table. See schema for details.
                    datarow = list()
                    datarow.append('null') # id
                    datarow.append(0) # title_report_id
                    datarow.append(self.title_type) # title_type
                    datarow.append(1) # access_type --> defaults to Controlled
                    datarow.append(2) # metric_type --> defaults to Total_Item_Requests
                    datarow.append(period)
                    datarow.append(period_total)
                    datarow.append(self._filename)
                    datarow.append(row_num)
                    csvwriter.writerow(datarow)
                    n += 1
                row_num += 1
